from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from loguru import logger

from .loader_registry import LoaderRegistry, DocumentContent


@LoaderRegistry.register([".xlsx", ".xls"])
class ExcelLoader:
    def load(self, file_path: Path) -> DocumentContent:
        wb = load_workbook(str(file_path), data_only=True)
        parts: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                parts.append(f"## Sheet: {sheet_name}\n\n" + "\n".join(rows))

        wb.close()
        full_text = "\n\n".join(parts)
        logger.info(f"Excel loaded: {file_path.name}, {len(full_text)} chars")
        return DocumentContent(text=full_text)
